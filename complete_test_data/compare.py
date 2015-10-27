#!/user/bin/python3
# -*- coding: utf-8 -*-
# compatible with Python 3.4.3

__author__="Bo Zhou"
__copyright__ = "Copyright 2015, The NRA project "
__credits__ = ["Bo Zhou"]
__license__ = "MIT"
__version__ = "1.0.0"
__maintainer__ = "Bo Zhou"
__email__ = "bzhou2@ualberta.ca"
__status__ = "Testing"

from openpyxl import Workbook
from openpyxl import load_workbook
import time

inFile = input("Please enter xlsx file name: ")
startTime = time.time()
inwb = load_workbook(filename='data_confirm_PA.xlsx')
sheetRange = inwb ['NRA Address']
row = 1
reference = []
while True:
    if sheetRange["A"+str(row)].value == None:
        break    
    item1 = sheetRange["A"+str(row)].value
    item2 = sheetRange["B"+str(row)].value
    item3 = sheetRange["C"+str(row)].value
    item4 = sheetRange["D"+str(row)].value
    item5 = sheetRange["E"+str(row)].value
    itemTuple = (item1,item2,item3,item4,item5)
    reference.append(itemTuple)
    row += 1
#inwb.save('data_confirm_PA.xlsx')

row = 1
cwb = load_workbook(filename=inFile)
sheetRange2 = cwb ['NRA Address']
compare = []
while True:
    if sheetRange2["A"+str(row)].value == None:
        break    
    item1 = sheetRange2["A"+str(row)].value
    item2 = sheetRange2["B"+str(row)].value
    item3 = sheetRange2["C"+str(row)].value
    item4 = sheetRange2["D"+str(row)].value
    item5 = sheetRange2["E"+str(row)].value
    itemTuple = (item1,item2,item3,item4,item5)
    compare.append(itemTuple)
    row += 1
#cwb.save(inFile)
bad = 0
for e in compare:
    if e not in reference:
        rownumber = compare.index(e)
        print ("on Row: "+str(rownumber+1)+" . Bad data: ",str(e[1]))
        bad += 1

if bad == 0 :
    print ("All data are found in reference")
print("Finish")
elapsedTime = time.time() - startTime
print ("time use: ", elapsedTime)